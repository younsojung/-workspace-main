"""xlsx 디코딩 + 모든 탭 추출"""
import json, base64
from pathlib import Path
import openpyxl

INPUT = "/Users/sojungyoun/.claude/projects/-Users-sojungyoun-Desktop-claude-workspace-main/6eafa475-1140-44ac-aeb5-1fc6b22ce00f/tool-results/mcp-claude_ai_Google_Drive-download_file_content-1779508349378.txt"
OUT_DIR = Path("/Users/sojungyoun/Desktop/claude/workspace-main/10-projects/younsojung-site-renewal/data/raw")
OUT_DIR.mkdir(parents=True, exist_ok=True)

with open(INPUT, encoding="utf-8") as f:
    payload = json.load(f)

print(f"[meta] title={payload.get('title')!r} mime={payload.get('mimeType')!r}")
raw = base64.b64decode(payload["content"])
xlsx_path = OUT_DIR / "form-responses-full.xlsx"
with open(xlsx_path, "wb") as f:
    f.write(raw)
print(f"[save] {xlsx_path}  ({len(raw):,} bytes)")

wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
print(f"\n[sheets] {len(wb.sheetnames)}")
for name in wb.sheetnames:
    ws = wb[name]
    # 행/열 추정 (read_only mode)
    rows = list(ws.iter_rows(values_only=True))
    n_rows = len(rows)
    n_cols = max((len(r) for r in rows), default=0)
    # 첫 행 (헤더 후보)
    header = rows[0] if rows else ()
    print(f"\n[sheet] {name!r}  rows={n_rows}  cols={n_cols}")
    print(f"  header preview: {[str(c)[:30] if c else '' for c in header[:14]]}")
    # 첫 데이터 행과 마지막 데이터 행
    if n_rows >= 2:
        print(f"  first row  : {[str(c)[:30] if c else '' for c in rows[1][:8]]}")
        print(f"  last row   : {[str(c)[:30] if c else '' for c in rows[-1][:8]]}")
