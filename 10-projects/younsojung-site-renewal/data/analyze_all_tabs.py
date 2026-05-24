"""3개 탭 통합 분석: 7년치 (2019-2026)
- 변하지 않은 핵심 vs 시대별 변화
- 연도별 카테고리 분포
"""
import re, csv
from pathlib import Path
from collections import Counter, defaultdict
from datetime import datetime
import openpyxl

XLSX = "/Users/sojungyoun/Desktop/claude/workspace-main/10-projects/younsojung-site-renewal/data/raw/form-responses-full.xlsx"
OUT_DIR = Path("/Users/sojungyoun/Desktop/claude/workspace-main/10-projects/younsojung-site-renewal/data")

wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
print(f"[load] sheets = {wb.sheetnames}")

# 각 탭별 컬럼 매핑 (탐색 후 수동 지정)
# 시트1: date=0, plan=2, wish=7  (header row 0)
# 과거1: date=0, plan=2, wish=7  (header row 0)
# 과거2: 헤더 없음. 모든 행을 데이터로. wish 컬럼 자동 탐지 필요.

def to_dt(v):
    if isinstance(v, datetime):
        return v
    if not v: return None
    s = str(v).strip()
    # ISO datetime first
    try:
        return datetime.fromisoformat(s.replace("Z", "").split(".")[0])
    except Exception:
        pass
    m = re.match(r"(\d{4})\.\s*(\d{1,2})\.\s*(\d{1,2})\s*(오전|오후)?\s*(\d{1,2}):(\d{2}):(\d{2})", s)
    if m:
        y,mo,d,ampm,hh,mi,ss = m.groups()
        hh = int(hh)
        if ampm=="오후" and hh<12: hh+=12
        elif ampm=="오전" and hh==12: hh=0
        return datetime(int(y),int(mo),int(d),hh,int(mi),int(ss))
    return None

# 통합 행 수집: {date, plan, wish}
all_entries = []
sheet_counts = {}

def has_header(first_row):
    """첫 셀이 '타임스탬프'인지로 헤더 유무 판단"""
    if not first_row: return False
    c0 = str(first_row[0]) if first_row[0] else ""
    return "타임스탬프" in c0

# wish 컬럼 자동 탐지 (헤더 없을 때): 가장 긴 한국어 텍스트를 담은 컬럼
def detect_wish_col(rows, date_col, max_check=200):
    if not rows: return None
    n_cols = max(len(r) for r in rows[:max_check])
    col_avg_len = {}
    col_korean_score = {}
    for col in range(n_cols):
        if col == date_col: continue
        lens = []
        korean = 0
        for r in rows[:max_check]:
            if col >= len(r): continue
            v = r[col]
            if v is None: continue
            s = str(v).strip()
            if not s: continue
            lens.append(len(s))
            if re.search(r"[가-힣]{5,}", s):
                korean += 1
        if lens:
            col_avg_len[col] = sum(lens)/len(lens)
            col_korean_score[col] = korean
    # 평균 길이 20자+ & 한글 비율 높은 컬럼 우선
    candidates = [(col, col_avg_len[col], col_korean_score[col]) for col in col_avg_len if col_avg_len[col] >= 20]
    candidates.sort(key=lambda x: (-x[1], -x[2]))
    return candidates[0][0] if candidates else None

for sname in wb.sheetnames:
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    if has_header(rows[0]):
        header = rows[0]
        data = rows[1:]
        # 헤더에서 wish 컬럼 식별
        wish_idx = None
        plan_idx = None
        for i, h in enumerate(header):
            hs = str(h) if h else ""
            if "나누고 싶은" in hs or "기대하는" in hs:
                wish_idx = i
            if "기간" in hs or "받아보" in hs:
                plan_idx = i
        date_idx = 0
    else:
        # 과거신청자2: 헤더 없음 — 수동 매핑 (확인 완료)
        data = rows
        date_idx = 0
        plan_idx = 4
        wish_idx = 9
    print(f"\n[{sname}] rows={len(data)}, date={date_idx}, plan={plan_idx}, wish={wish_idx}")
    valid = 0
    for r in data:
        if not r: continue
        if len(r) <= max(filter(None, [date_idx, wish_idx, plan_idx or 0])):
            continue
        dt = to_dt(r[date_idx])
        if dt is None: continue
        plan = str(r[plan_idx]).strip() if plan_idx is not None and r[plan_idx] else ""
        wish = str(r[wish_idx]).strip() if wish_idx is not None and r[wish_idx] else ""
        all_entries.append({"sheet": sname, "dt": dt, "plan": plan, "wish": wish})
        valid += 1
    sheet_counts[sname] = valid

print(f"\n[total] entries with valid date = {len(all_entries):,}")
for s, c in sheet_counts.items():
    print(f"  - {s}: {c:,}")

# 시간 범위 출력
dts = [e["dt"] for e in all_entries]
print(f"[range] {min(dts)} ~ {max(dts)}")

# 답변 있는 것만
wish_entries = [e for e in all_entries if e["wish"] and len(e["wish"]) >= 4]
print(f"[wish] non-empty answers = {len(wish_entries):,} ({len(wish_entries)/len(all_entries)*100:.1f}%)")

# === 카테고리 (동일 사전) ===
CATEGORIES = {
    "AI/도구": ["AI","ai","인공지능","챗gpt","ChatGPT","chatGPT","GPT","claude","클로드","프롬프트","자동화","툴","도구"],
    "일/커리어/사업": ["일","커리어","직장","직업","이직","퇴사","회사","비즈니스","비지니스","사업","창업","프리랜서","1인기업","워라밸","성과"],
    "돈/투자/재테크": ["돈","수익","재테크","투자","자산","부동산","경제","재무","수익화","현금흐름","월급","연봉","파이프라인","패시브"],
    "관계/부모/자녀": ["관계","인간관계","사람","부모","엄마","아빠","자녀","아이","딸","아들","남편","아내","배우자","가족","친구","동료"],
    "성장/마인드": ["성장","마인드","마인드셋","태도","용기","자존감","자신감","감정","스트레스","불안","두려움","무기력","번아웃"],
    "습관/시간/루틴": ["습관","루틴","시간","아침","새벽","미라클","계획","실행","꾸준","지속","리추얼"],
    "글쓰기/콘텐츠": ["글","글쓰기","라이팅","콘텐츠","퍼스널브랜드","퍼스널 브랜드","브랜딩","유튜브","인스타","sns","SNS","블로그"],
    "공부/학습/독서": ["공부","학습","독서","책","강의","배움","배우"],
    "진로/미래/방향": ["진로","미래","방향","방향성","꿈","비전","목표","what to do","뭘 해야"],
    "마케팅/판매": ["마케팅","세일즈","판매","고객","고객언어","타겟","포지셔닝","전환"],
    "건강/운동": ["건강","운동","다이어트","체력","수면","잠"],
    "영성/명상": ["명상","영성","기도","감사","내면","영혼","에너지"],
    "육아/교육": ["육아","교육","양육","엄마로","워킹맘","초등","유아","사춘기"],
    "협상/소통": ["협상","대화","소통","표현","말하기","발표","거절"],
    "윤소정 본인/생각": ["소정님","소정 님","선생님","소정쌤","작가님","윤소정","생각구독","생각 구독","이야기","관점","통찰","지혜"],
    "결단/선택": ["결단","결정","선택","기준","판단"],
    "환경설계": ["환경","환경설계","환경 설계","시스템","구조"],
}

def categorize(text):
    matched = set()
    for cat, kws in CATEGORIES.items():
        if any(kw in text for kw in kws):
            matched.add(cat)
    return matched

# 연도별 분포
by_year_cat = defaultdict(Counter)
by_year_total = Counter()
by_year_wish = Counter()
for e in wish_entries:
    yr = e["dt"].year
    by_year_wish[yr] += 1
    for cat in categorize(e["wish"]):
        by_year_cat[yr][cat] += 1
for e in all_entries:
    by_year_total[e["dt"].year] += 1

# 전체 카테고리
total_cat = Counter()
for e in wish_entries:
    for cat in categorize(e["wish"]):
        total_cat[cat] += 1

# === 출력 ===
out = OUT_DIR / "7yr-evolution.md"
with open(out, "w", encoding="utf-8") as f:
    f.write("# 7년치 통합 분석 (2019-10 ~ 2026-05)\n\n")
    f.write(f"- 전체 신청: **{len(all_entries):,}건**\n")
    f.write(f"- 자유서술 답변: **{len(wish_entries):,}건** ({len(wish_entries)/len(all_entries)*100:.1f}%)\n")
    f.write(f"- 기간: {min(dts).date()} ~ {max(dts).date()}\n\n")
    f.write("## 시트별\n")
    for s, c in sheet_counts.items():
        f.write(f"- {s}: {c:,}건\n")
    f.write("\n## 연도별 신청 수\n\n")
    f.write("| 연도 | 신청 | 답변 작성 | 답변률 |\n|---|---:|---:|---:|\n")
    for yr in sorted(by_year_total):
        t = by_year_total[yr]; w = by_year_wish[yr]
        f.write(f"| {yr} | {t:,} | {w:,} | {w/t*100:.1f}% |\n")

    f.write("\n---\n\n## 🌳 7년 누적 카테고리 (변하지 않은 본질)\n\n")
    f.write("| 카테고리 | 답변 수 | 비중 (전체 답변 기준) |\n|---|---:|---:|\n")
    for cat, cnt in total_cat.most_common():
        f.write(f"| {cat} | {cnt:,} | {cnt/len(wish_entries)*100:.1f}% |\n")

    f.write("\n---\n\n## 📈 연도별 카테고리 비중 (% — 그해 답변 대비)\n\n")
    years = sorted(by_year_cat)
    cats_top = [c for c,_ in total_cat.most_common(10)]
    f.write("| 카테고리 | " + " | ".join(str(y) for y in years) + " |\n")
    f.write("|---|" + "|".join(["---:"]*len(years)) + "|\n")
    for cat in cats_top:
        row = [cat]
        for yr in years:
            total = by_year_wish[yr]
            pct = by_year_cat[yr][cat] / total * 100 if total else 0
            row.append(f"{pct:.1f}%")
        f.write("| " + " | ".join(row) + " |\n")

    # 가장 안정적인 카테고리 (편차 최소) vs 시대성 (편차 최대)
    f.write("\n---\n\n## 🎯 본질 vs 시대성\n\n")
    stability = []
    for cat in cats_top:
        pcts = []
        for yr in years:
            total = by_year_wish[yr]
            if total >= 50:  # 최소 50개 이상 답변 있는 해만
                pcts.append(by_year_cat[yr][cat] / total * 100)
        if len(pcts) >= 3:
            mean = sum(pcts)/len(pcts)
            var = sum((p-mean)**2 for p in pcts)/len(pcts)
            stability.append((cat, mean, var**0.5))
    stability.sort(key=lambda x: x[2])
    f.write("### 🌳 가장 안정적인 카테고리 (= 변하지 않은 본질)\n\n")
    f.write("| 카테고리 | 평균 비중 | 표준편차 |\n|---|---:|---:|\n")
    for cat, mean, sd in stability[:5]:
        f.write(f"| {cat} | {mean:.1f}% | ±{sd:.1f}%p |\n")
    f.write("\n### 🌀 가장 변동성 큰 카테고리 (= 시대성)\n\n")
    f.write("| 카테고리 | 평균 비중 | 표준편차 |\n|---|---:|---:|\n")
    for cat, mean, sd in sorted(stability, key=lambda x: -x[2])[:5]:
        f.write(f"| {cat} | {mean:.1f}% | ±{sd:.1f}%p |\n")

print(f"\n[save] {out}")

# 통합 wish CSV도 저장
import csv as _csv
csv_path = OUT_DIR / "raw" / "all-wishes-7yr.csv"
with open(csv_path, "w", encoding="utf-8", newline="") as f:
    w = _csv.writer(f)
    w.writerow(["sheet","datetime","year","month","plan","wish"])
    for e in wish_entries:
        w.writerow([e["sheet"], e["dt"].isoformat(), e["dt"].year, e["dt"].strftime("%Y-%m"), e["plan"], e["wish"]])
print(f"[save] {csv_path}")
print("\n[done]")
