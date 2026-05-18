#!/usr/bin/env python3
"""Excel/CSV to UTF-8 CSV converter for Claude Code.

Usage:
  python excel-to-csv.py <file_or_folder> [options]

Options:
  --info              Show sheet list, preview, and complexity analysis
  --sheet <name>      Convert specific sheet only
  --output <path>     Output directory (default: same as source)
  --all               Convert all sheets without confirmation
  --encoding <enc>    Force source encoding (for CSV input)
  --flatten-headers   Flatten multi-row headers into "Parent_Child" format
  --header-rows N     Number of header rows (default: auto-detect)
  --skip-rows N       Skip top N rows (metadata)
"""

import argparse
import csv
import os
import re
import sys
from datetime import datetime, date, time
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl>=3.1.0", file=sys.stderr)
    sys.exit(1)


def format_cell(value):
    """Format a cell value for CSV output. Handles datetime objects cleanly."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return str(value)


def sanitize_filename(name: str) -> str:
    """Sanitize filename: keep Korean, replace spaces, remove special chars."""
    name = name.replace(" ", "_")
    name = re.sub(r'[<>:"/\\|?*]', "", name)
    return name


def detect_encoding(file_path: str) -> str:
    """Detect file encoding by trying common Korean encodings."""
    encodings = ["utf-8", "utf-8-sig", "euc-kr", "cp949", "utf-16", "utf-16-le"]
    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc) as f:
                f.read(4096)
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue
    return "utf-8"


def find_header_row(ws) -> int:
    """Find the first non-empty row (0-indexed) as header."""
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if any(cell is not None and str(cell).strip() for cell in row):
            return idx
    return 0


def get_max_col(ws) -> int:
    """Find the last non-empty column."""
    max_col = 0
    for row in ws.iter_rows(values_only=False):
        for cell in reversed(row):
            if cell.value is not None and str(cell.value).strip():
                max_col = max(max_col, cell.column)
                break
    return max_col


def get_max_row(ws, header_row: int) -> int:
    """Find the last non-empty row after header."""
    max_row = header_row
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx >= header_row and any(cell is not None and str(cell).strip() for cell in row):
            max_row = idx
    return max_row


def handle_merged_cells(ws):
    """Fill merged cell regions with the top-left value."""
    for merge_range in list(ws.merged_cells.ranges):
        min_row = merge_range.min_row
        min_col = merge_range.min_col
        value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merge_range))
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                ws.cell(row=row, column=col).value = value


# --- New: Header Detection & Flattening ---

def detect_header_rows(ws) -> int:
    """Auto-detect number of header rows.

    Heuristics:
    - Merged cells concentrated at top rows
    - Consecutive rows with text+empty alternating pattern
    - Type difference between header candidates (text) and data rows (numbers)
    """
    max_col = get_max_col(ws)
    if max_col == 0:
        return 1

    first_data_row = find_header_row(ws)

    # Collect merged cell info in top rows
    merged_rows = set()
    for merge_range in ws.merged_cells.ranges:
        if merge_range.min_row <= first_data_row + 5:
            for r in range(merge_range.min_row, merge_range.max_row + 1):
                merged_rows.add(r)

    # Analyze top rows: check if they look like headers (mostly text) vs data (has numbers)
    candidate_rows = []
    for row_idx in range(first_data_row + 1, min(first_data_row + 6, ws.max_row + 1)):
        row_values = []
        for col_idx in range(1, max_col + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            row_values.append(val)

        text_count = sum(1 for v in row_values if v is not None and isinstance(v, str) and v.strip())
        num_count = sum(1 for v in row_values if isinstance(v, (int, float)))
        empty_count = sum(1 for v in row_values if v is None or (isinstance(v, str) and not v.strip()))

        candidate_rows.append({
            "row": row_idx,
            "text": text_count,
            "num": num_count,
            "empty": empty_count,
            "has_merge": row_idx in merged_rows,
        })

    if not candidate_rows:
        return 1

    # Find transition point: where numbers start appearing significantly
    for i, info in enumerate(candidate_rows):
        total_filled = info["text"] + info["num"]
        if total_filled == 0:
            continue
        num_ratio = info["num"] / total_filled if total_filled > 0 else 0
        # If this row has >40% numbers, it's likely data, not header
        if num_ratio > 0.4 and i > 0:
            return i
        # If previous rows had merges but this one doesn't, and it has numbers
        if i > 0 and candidate_rows[i - 1]["has_merge"] and not info["has_merge"] and num_ratio > 0.2:
            return i

    # Default: check if first 2 rows both look like text headers
    if len(candidate_rows) >= 2:
        r1, r2 = candidate_rows[0], candidate_rows[1]
        if r1["text"] > 0 and r2["text"] > 0 and r1["num"] == 0 and r2["num"] == 0:
            # Both look like text headers, check if row 3 has numbers
            if len(candidate_rows) >= 3 and candidate_rows[2]["num"] > 0:
                return 2

    return 1


def flatten_multi_headers(ws, num_header_rows, start_row, max_col):
    """Flatten multi-row headers into 'Parent_Child' format.

    Must be called AFTER handle_merged_cells.
    Returns list of flattened header strings.
    """
    headers_by_row = []
    for r in range(start_row + 1, start_row + num_header_rows + 1):
        row_headers = []
        for c in range(1, max_col + 1):
            val = ws.cell(row=r, column=c).value
            row_headers.append(str(val).strip() if val is not None else "")
        headers_by_row.append(row_headers)

    flattened = []
    for col_idx in range(max_col):
        parts = []
        for row_headers in headers_by_row:
            if col_idx < len(row_headers):
                val = row_headers[col_idx]
                if val:
                    parts.append(val)

        # Remove duplicates (e.g., "Seoul" + "Seoul" -> "Seoul")
        deduped = []
        for p in parts:
            if not deduped or deduped[-1] != p:
                deduped.append(p)

        flattened.append("_".join(deduped) if deduped else f"col_{col_idx + 1}")

    return flattened


# --- New: Complexity Analysis ---

SUBTOTAL_KEYWORDS = ["소계", "합계", "계", "총계", "총합", "subtotal", "total", "sum"]
MONTH_PATTERNS = [
    re.compile(r'^\d{1,2}월$'),
    re.compile(r'^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)', re.I),
    re.compile(r'^\d{4}[-/.]\d{1,2}$'),
    re.compile(r'^[qQ][1-4]$'),
    re.compile(r'^\d{1,2}분기$'),
]
NUMBER_FORMAT_RE = re.compile(r'^[\s]*[$₩¥€£]?[\s]*[\d,]+\.?\d*[\s]*%?[\s]*$')
DATE_FORMAT_RE = re.compile(r'^\d{4}[-/.]\d{1,2}[-/.]\d{1,2}$|^\d{8}$|^\d{4}년')


def analyze_complexity(ws):
    """Analyze worksheet complexity and print suggestions.

    Called during --info. Detects:
    - Multi-row headers
    - Metadata rows before header
    - Subtotal rows (suggests csv-clean)
    - Text-formatted numbers (suggests csv-clean)
    - Crosstab structure (suggests csv-clean)
    - Mixed date formats (suggests csv-clean)
    """
    handle_merged_cells(ws)
    max_col = get_max_col(ws)
    if max_col == 0:
        return

    header_start = find_header_row(ws)
    max_row = get_max_row(ws, header_start)
    detected_header_rows = detect_header_rows(ws)

    # Find the "real" header row: the row right before data starts
    # Data rows have significant numeric content; header rows are mostly text
    real_header_row = header_start  # 0-indexed
    data_start_candidate = header_start + detected_header_rows

    # Scan from header_start to find the last all-text row before data
    for row_idx in range(header_start + 1, min(header_start + 8, ws.max_row + 1)):
        row_values = []
        for col_idx in range(1, max_col + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            row_values.append(val)

        filled = [v for v in row_values if v is not None and str(v).strip()]
        nums = [v for v in filled if isinstance(v, (int, float))]
        texts = [v for v in filled if isinstance(v, str) and v.strip()]

        # If this row has numbers and filled cells, it's likely data
        if len(nums) > 0 and len(filled) > max_col * 0.3:
            # The previous row with mostly text is the real header
            real_header_row = row_idx - 2  # 0-indexed (row before this one)
            data_start_candidate = row_idx - 1  # 0-indexed
            break

    # Calculate recommended skip_rows: everything before the real column headers
    # real_header_row is the 0-indexed row of the actual column headers
    recommended_skip = max(0, data_start_candidate - detected_header_rows)

    print("--- Complexity Analysis ---")
    found = False

    # 1. Multi-row header detection
    if detected_header_rows > 1:
        found = True
        end_row = header_start + detected_header_rows
        print(f"[MULTI_HEADER] {detected_header_rows}-row header detected (rows {header_start + 1}-{end_row}). Suggested: --flatten-headers")

    # 2. Metadata rows before header
    if header_start > 0:
        found = True
        print(f"[METADATA_SKIP] {header_start} metadata rows before header. Suggested: --skip-rows {header_start}")

    # 3. Subtotal rows (csv-clean territory)
    data_start = header_start + detected_header_rows + 1
    subtotal_count = 0
    for row_idx in range(data_start, max_row + 2):
        for col_idx in range(1, min(4, max_col + 1)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                val_lower = str(val).strip().lower()
                if any(kw in val_lower for kw in SUBTOTAL_KEYWORDS):
                    subtotal_count += 1
                    break

    if subtotal_count > 0:
        found = True
        print(f"[SUBTOTAL_ROWS] {subtotal_count} rows with subtotals. -> Use csv-clean --remove-subtotals after conversion")

    # 4. Text-formatted numbers (csv-clean territory)
    num_cols_with_text = 0
    for col_idx in range(1, max_col + 1):
        sample_count = 0
        for row_idx in range(data_start, min(data_start + 20, max_row + 2)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                s = str(val).strip()
                if ("," in s or "$" in s or "₩" in s or "%" in s) and any(c.isdigit() for c in s):
                    sample_count += 1
        if sample_count >= 3:
            num_cols_with_text += 1

    if num_cols_with_text > 0:
        found = True
        print(f"[TEXT_NUMBERS] {num_cols_with_text} columns with formatted numbers. -> Use csv-clean --clean-numbers after conversion")

    # 5. Crosstab detection (csv-clean territory)
    headers = []
    header_row_num = header_start + 1
    for col_idx in range(1, max_col + 1):
        val = ws.cell(row=header_row_num, column=col_idx).value
        if val is not None:
            headers.append(str(val).strip())

    month_matches = 0
    for h in headers:
        for p in MONTH_PATTERNS:
            if p.match(h):
                month_matches += 1
                break

    if month_matches >= 3:
        found = True
        matched_headers = []
        for h in headers:
            for p in MONTH_PATTERNS:
                if p.match(h):
                    matched_headers.append(h)
                    break
        print(f"[CROSSTAB] Headers contain month/quarter pattern: {', '.join(matched_headers[:6])}. -> Use csv-clean --unpivot after conversion")

    # 6. Mixed date formats (csv-clean territory)
    date_cols = 0
    for col_idx in range(1, max_col + 1):
        date_count = 0
        for row_idx in range(data_start, min(data_start + 20, max_row + 2)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None and DATE_FORMAT_RE.match(str(val).strip()):
                date_count += 1
        if date_count >= 3:
            date_cols += 1

    if date_cols > 0:
        found = True
        print(f"[DATE_FORMATS] Mixed date formats detected in {date_cols} column(s). -> Use csv-clean --normalize-dates after conversion")

    if not found:
        print("No complexity issues detected. Standard conversion should work fine.")

    # Recommended command
    opts = []
    if recommended_skip > 0:
        opts.append(f"--skip-rows {recommended_skip}")
    if detected_header_rows > 1:
        opts.append("--flatten-headers")
    opts.append("--all")

    if opts:
        print()
        print(f"[RECOMMENDED] python excel-to-csv.py <file> {' '.join(opts)}")
        if subtotal_count > 0 or num_cols_with_text > 0 or date_cols > 0:
            clean_opts = []
            if subtotal_count > 0:
                clean_opts.append("--remove-subtotals")
            if num_cols_with_text > 0:
                clean_opts.append("--clean-numbers")
            if date_cols > 0:
                clean_opts.append("--normalize-dates")
            print(f"[RECOMMENDED] Then: python csv-clean.py <output.csv> {' '.join(clean_opts)}")
    print()


def sheet_info(ws, sheet_name: str) -> dict:
    """Get sheet metadata."""
    header_row = find_header_row(ws)
    max_col = get_max_col(ws)
    max_row = get_max_row(ws, header_row)
    data_rows = max_row - header_row  # excluding header

    if max_col == 0 or data_rows <= 0:
        return {"name": sheet_name, "rows": 0, "cols": 0, "headers": [], "empty": True}

    headers = []
    for cell in list(ws.iter_rows(min_row=header_row + 1, max_row=header_row + 1,
                                   min_col=1, max_col=max_col, values_only=True))[0]:
        headers.append(format_cell(cell))

    preview_rows = []
    row_iter = ws.iter_rows(min_row=header_row + 2, max_row=min(header_row + 6, max_row + 1),
                            min_col=1, max_col=max_col, values_only=True)
    for row in row_iter:
        preview_rows.append([format_cell(c) for c in row])

    return {
        "name": sheet_name,
        "rows": data_rows,
        "cols": max_col,
        "headers": headers,
        "preview": preview_rows,
        "empty": False,
    }


def print_info(file_path: str):
    """Print sheet information and complexity analysis for an Excel file."""
    wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
    print(f"File: {file_path}")
    print(f"Sheets: {len(wb.sheetnames)}")
    print("-" * 60)

    for name in wb.sheetnames:
        ws = wb[name]
        info = sheet_info(ws, name)
        if info["empty"]:
            print(f"  [{name}] (empty - skipped)")
            continue
        print(f"  [{name}] {info['rows']} rows x {info['cols']} cols")
        print(f"    Headers: {', '.join(info['headers'][:10])}")
        if len(info["headers"]) > 10:
            print(f"    ... and {len(info['headers']) - 10} more columns")
        if info.get("preview"):
            print(f"    Preview (first {len(info['preview'])} rows):")
            for row in info["preview"][:3]:
                truncated = [v[:30] + "..." if len(v) > 30 else v for v in row[:6]]
                print(f"      {' | '.join(truncated)}")
        print()

    # Complexity analysis for each non-empty sheet
    for name in wb.sheetnames:
        ws = wb[name]
        max_col = get_max_col(ws)
        if max_col == 0:
            continue
        print(f"[{name}]")
        analyze_complexity(ws)

    wb.close()


def convert_sheet_to_csv(ws, output_path: str, sheet_name: str,
                         skip_rows: int = 0, flatten: bool = False,
                         header_rows_override: int = None) -> dict:
    """Convert a single worksheet to CSV. Returns metadata.

    Pipeline: merged cells -> skip rows -> header processing -> data -> CSV
    """
    handle_merged_cells(ws)
    header_row = find_header_row(ws)
    max_col = get_max_col(ws)
    max_row = get_max_row(ws, header_row)

    # Apply skip_rows
    effective_start = header_row + skip_rows

    if max_col == 0:
        return {"sheet": sheet_name, "rows": 0, "skipped": True}

    # Determine header rows count
    if flatten:
        if header_rows_override is not None:
            num_header_rows = header_rows_override
        else:
            num_header_rows = detect_header_rows(ws)
            # Re-detect relative to effective_start if skip applied
            if skip_rows > 0:
                # Simple re-detection: just use the override or default
                num_header_rows = max(num_header_rows, 1)
    else:
        num_header_rows = 1

    data_start_row = effective_start + num_header_rows  # 0-indexed
    data_rows = max_row - data_start_row

    if data_rows <= 0:
        return {"sheet": sheet_name, "rows": 0, "skipped": True}

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        # Write header
        if flatten and num_header_rows > 1:
            headers = flatten_multi_headers(ws, num_header_rows, effective_start, max_col)
            writer.writerow(headers)
        else:
            # Single header row
            header_ws_row = effective_start + 1  # 1-indexed for openpyxl
            row_data = list(ws.iter_rows(min_row=header_ws_row, max_row=header_ws_row,
                                          min_col=1, max_col=max_col, values_only=True))[0]
            writer.writerow([str(c) if c is not None else "" for c in row_data])

        # Write data rows
        data_ws_start = effective_start + num_header_rows + 1  # 1-indexed
        for row in ws.iter_rows(min_row=data_ws_start, max_row=max_row + 1,
                                min_col=1, max_col=max_col, values_only=True):
            writer.writerow([format_cell(c) for c in row])

    return {"sheet": sheet_name, "rows": data_rows, "path": output_path, "skipped": False}


def convert_excel(file_path: str, output_dir: str, sheet_name: str = None,
                  convert_all: bool = False, skip_rows: int = 0,
                  flatten: bool = False, header_rows: int = None):
    """Convert Excel file to CSV(s)."""
    wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
    stem = sanitize_filename(Path(file_path).stem)
    results = []

    sheets_to_convert = []
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"ERROR: Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}", file=sys.stderr)
            wb.close()
            sys.exit(1)
        sheets_to_convert = [sheet_name]
    else:
        sheets_to_convert = wb.sheetnames

    # Determine naming: single sheet = simple name, multi = with sheet suffix
    use_suffix = len(sheets_to_convert) > 1

    for name in sheets_to_convert:
        ws = wb[name]
        if use_suffix:
            csv_name = f"{stem}_{sanitize_filename(name)}.csv"
        else:
            csv_name = f"{stem}.csv"

        out_path = os.path.join(output_dir, csv_name)
        result = convert_sheet_to_csv(ws, out_path, name,
                                      skip_rows=skip_rows,
                                      flatten=flatten,
                                      header_rows_override=header_rows)
        results.append(result)

    wb.close()

    # Print results
    print(f"Converted: {Path(file_path).name}")
    converted_count = 0
    total_rows = 0
    for r in results:
        if r.get("skipped"):
            print(f"  [{r['sheet']}] skipped (empty)")
        else:
            print(f"  [{r['sheet']}] -> {Path(r['path']).name} ({r['rows']} rows)")
            converted_count += 1
            total_rows += r["rows"]
    if flatten:
        print(f"  (headers flattened: multi-row -> single row)")
    if skip_rows > 0:
        print(f"  (skipped {skip_rows} metadata rows)")
    print(f"Total: {converted_count} file(s), {total_rows} rows")
    return results


def convert_csv_encoding(file_path: str, output_dir: str, source_encoding: str = None):
    """Convert a CSV file from detected/specified encoding to UTF-8."""
    if source_encoding is None:
        source_encoding = detect_encoding(file_path)

    stem = sanitize_filename(Path(file_path).stem)
    out_path = os.path.join(output_dir, f"{stem}_utf8.csv")

    with open(file_path, "r", encoding=source_encoding) as infile:
        content = infile.read()

    with open(out_path, "w", encoding="utf-8", newline="") as outfile:
        outfile.write(content)

    # Count rows
    row_count = content.count("\n")
    print(f"Converted: {Path(file_path).name}")
    print(f"  Encoding: {source_encoding} -> UTF-8")
    print(f"  Output: {Path(out_path).name} ({row_count} rows)")
    return [{"sheet": "csv", "rows": row_count, "path": out_path, "skipped": False}]


def process_folder(folder_path: str, output_dir: str, convert_all: bool = False,
                   skip_rows: int = 0, flatten: bool = False, header_rows: int = None):
    """Process all Excel files in a folder."""
    folder = Path(folder_path)
    excel_files = sorted(list(folder.glob("*.xlsx")) + list(folder.glob("*.xls")) + list(folder.glob("*.XLSX")))

    if not excel_files:
        print(f"No Excel files found in {folder_path}")
        return

    print(f"Found {len(excel_files)} Excel file(s) in {folder_path}")
    print("-" * 60)

    all_results = []
    for fp in excel_files:
        results = convert_excel(str(fp), output_dir, convert_all=convert_all,
                                skip_rows=skip_rows, flatten=flatten, header_rows=header_rows)
        all_results.extend(results)
        print()

    total_files = sum(1 for r in all_results if not r.get("skipped"))
    total_rows = sum(r.get("rows", 0) for r in all_results if not r.get("skipped"))
    print("=" * 60)
    print(f"Grand total: {len(excel_files)} Excel file(s) -> {total_files} CSV file(s), {total_rows} rows")


def main():
    parser = argparse.ArgumentParser(description="Excel/CSV to UTF-8 CSV converter")
    parser.add_argument("path", help="Excel file or folder path")
    parser.add_argument("--info", action="store_true", help="Show sheet info and complexity analysis")
    parser.add_argument("--sheet", help="Convert specific sheet only")
    parser.add_argument("--output", help="Output directory")
    parser.add_argument("--all", action="store_true", dest="convert_all", help="Convert all sheets")
    parser.add_argument("--encoding", help="Force source encoding (for CSV files)")
    parser.add_argument("--flatten-headers", action="store_true", help="Flatten multi-row headers")
    parser.add_argument("--header-rows", type=int, help="Number of header rows (default: auto-detect)")
    parser.add_argument("--skip-rows", type=int, default=0, help="Skip top N rows (metadata)")

    args = parser.parse_args()
    target = os.path.abspath(args.path)

    if not os.path.exists(target):
        print(f"ERROR: Path not found: {target}", file=sys.stderr)
        sys.exit(1)

    # Determine output directory
    if args.output:
        output_dir = os.path.abspath(args.output)
        os.makedirs(output_dir, exist_ok=True)
    elif os.path.isfile(target):
        output_dir = os.path.dirname(target)
    else:
        output_dir = target

    # Folder mode
    if os.path.isdir(target):
        if args.info:
            for fp in sorted(Path(target).glob("*.xlsx")) + sorted(Path(target).glob("*.XLSX")):
                print_info(str(fp))
                print()
        else:
            process_folder(target, output_dir, args.convert_all,
                           skip_rows=args.skip_rows,
                           flatten=args.flatten_headers,
                           header_rows=args.header_rows)
        return

    # Single file mode
    ext = Path(target).suffix.lower()

    if ext in (".xlsx", ".xls"):
        if args.info:
            print_info(target)
        else:
            convert_excel(target, output_dir, args.sheet, args.convert_all,
                          skip_rows=args.skip_rows,
                          flatten=args.flatten_headers,
                          header_rows=args.header_rows)
    elif ext == ".csv":
        convert_csv_encoding(target, output_dir, args.encoding)
    else:
        print(f"ERROR: Unsupported file type: {ext}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
